import openpyxl, base64
from django.shortcuts import HttpResponse
from django.db.models import QuerySet
from django.core.exceptions import FieldError


class QuerySetToExcel:
    
    def __init__(self, queryset, schema) -> None:
        self.queryset = queryset
        self.schema = schema
        self.wb = self.to_excel(queryset, schema)

    def put_to_response(self, name="export", adjust_col_width=True):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{name}.xlsx"'

        if adjust_col_width:
            self.set_column_width(self.wb)

        self.wb.save(response)
        return response


    def export_in_file(self, name="export", adjust_col_width=True):
        pass



    def to_excel(self, queryset:QuerySet, schema,  rtl=True) -> openpyxl.Workbook:
        # Create a new Workbook
        wb = openpyxl.Workbook()
        # Select the active worksheet
        ws = wb.active
        if rtl:
            ws.sheet_view.rightToLeft = True
        
        # Write column headers using the keys of the schema
        for idx, column_name in enumerate(schema.keys(), start=1):
            ws.cell(row=1, column=idx, value=column_name)
        
        # Iterate over queryset and write data to Excel
        for row_idx, obj in enumerate(queryset.iterator(chunk_size=1000), start=2):
            for col_idx, field_name in enumerate(schema.values(), start=1):
                # Check if the field_name contains '.' indicating a relation or JSONField item
                if callable(field_name):
                    value = field_name(obj)
                elif '.' in field_name:
                    # Split the field name by '.' to handle relations or JSONField items
                    field_parts = field_name.split('.')
                    value = obj
                    # Traverse through the nested relations or JSONField items
                    for part in field_parts:
                        if isinstance(value, dict):
                            # If value is a dictionary (JSONField item), get the value by key
                            value = value.get(part)
                        elif hasattr(value, part):
                            # If value is an object (related model), get the attribute value
                            value = getattr(value, part)
                        else:
                            value = None
                            break
                else:
                    # If no '.' found, directly get the attribute value
                    value = getattr(obj, field_name, None)
                

                if callable(value):
                    value = value()

                # Write the value to the appropriate cell
                ws.cell(row=row_idx, column=col_idx, value=value)
   
        return wb


    def set_column_width(self, wb):
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Adjust the width to provide some padding
                ws.column_dimensions[column].width = adjusted_width
